"""Export today's NBA predictions to the OneDrive folder.

Generates:
  - YYYY-MM-DD/GBSV_NBA_Predictions_HHMM.csv
  - YYYY-MM-DD/GBSV_NBA_Predictions_HHMM.html
  - YYYY-MM-DD/GBSV_NBA_Predictions_HHMM.json
  - YYYY-MM-DD/GBSV_NBA_Predictions_HHMM.txt
  - YYYY-MM-DD/GBSV_NBA_Predictions_HHMM.xlsx
  - latest/GBSV_NBA_Predictions.csv
  - latest/GBSV_NBA_Predictions.html
  - latest/GBSV_NBA_Predictions.json
  - latest/GBSV_NBA_Predictions.txt
  - latest/GBSV_NBA_Predictions.xlsx
    - nba_slate_YYYY-MM-DD.csv

This script uses the repo env contract and active profile selection.
Run scripts/setup-env.ps1 first to sync/select the desired profile.
"""
from __future__ import annotations

import asyncio
import json
import os
import re
import sys
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any, cast
from zoneinfo import ZoneInfo

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

_CST = ZoneInfo("US/Central")


def _env_value(key: str, default: str = "") -> str:
    from src.config import load_selected_env_values

    raw = os.getenv(key)
    if raw is not None and raw.strip():
        return raw.strip()

    selected = load_selected_env_values().get(key)
    if selected is not None and str(selected).strip():
        return str(selected).strip()

    return default


def _env_bool(key: str, default: bool) -> bool:
    value = _env_value(key)
    if not value:
        return default
    return value.lower() in {"1", "true", "yes", "on"}


def _resolve_export_root() -> Path:
    configured = _env_value("ONEDRIVE_EXPORT_ROOT")
    if configured:
        return Path(configured)

    one_drive_base = (
        os.getenv("ONEDRIVECOMMERCIAL")
        or os.getenv("ONEDRIVE")
        or os.getenv("ONEDRIVECONSUMER")
    )
    if one_drive_base:
        return (
            Path(one_drive_base)
            / "Early Stage Sport Ventures - Documents"
            / "NBA - Green Bier Sports"
        )

    raise RuntimeError(
        "Set ONEDRIVE_EXPORT_ROOT or configure a Windows OneDrive environment before running export_onedrive.py."
    )


def _resolve_export_database_url() -> tuple[str, bool]:
    from src.config import get_settings, resolve_database_url, resolve_settings_env_file

    settings = get_settings()
    database_url = resolve_database_url().strip()
    if not database_url:
        raise RuntimeError(
            "DATABASE_URL is not configured. Run scripts/setup-env.ps1 to sync/select the correct profile before export_onedrive.py."
        )

    allow_local_db = _env_bool("EXPORT_ALLOW_LOCAL_DB", False)
    if not allow_local_db and ("localhost" in database_url or "127.0.0.1" in database_url):
        raise RuntimeError(
            "export_onedrive.py refuses to use a local DATABASE_URL. "
            f"Select an Azure-backed profile with scripts/setup-env.ps1. Current env file: {resolve_settings_env_file()}"
        )

    return database_url, settings.db_ssl


async def main() -> None:
    from sqlalchemy import select
    from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
    from sqlalchemy.orm import selectinload

    from src.db.models import Game, Prediction, Team
    from src.notifications.teams import (
        _team_record,
        build_html_slate,
        build_slate_csv,
        extract_picks,
    )

    database_url, db_ssl = _resolve_export_database_url()
    clean_database_url = re.sub(r"[?&]ssl(?:mode)?=[^&]*", "", database_url).rstrip("?&")
    ssl_connect_arg = "require" if db_ssl else False
    one_drive_root = _resolve_export_root()
    engine = create_async_engine(
        clean_database_url,
        echo=False,
        connect_args={"ssl": ssl_connect_arg},
    )
    async_session = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    day_now_cst = datetime.now(_CST)
    day_start_cst = day_now_cst.replace(hour=0, minute=0, second=0, microsecond=0)
    day_end_cst = day_start_cst + timedelta(days=1)
    today = day_start_cst.date()
    day_start = day_start_cst.astimezone(UTC).replace(tzinfo=None)
    day_end = day_end_cst.astimezone(UTC).replace(tzinfo=None)

    async with async_session() as db:
        # Load games + teams first, then predictions
        game_result = await db.execute(
            select(Game)
            .options(
                selectinload(Game.home_team).selectinload(Team.season_stats),
                selectinload(Game.away_team).selectinload(Team.season_stats),
            )
            .where(
                Game.commence_time >= day_start,
                Game.commence_time < day_end,
            )
            .order_by(Game.commence_time)
        )
        games_by_id = {g.id: g for g in game_result.scalars().all()}

        if not games_by_id:
            print(f"No games found for {today}. Exiting.")
            await engine.dispose()
            return

        pred_result = await db.execute(
            select(Prediction)
            .where(Prediction.game_id.in_(list(games_by_id.keys())))
            .order_by(Prediction.game_id)
        )
        preds = pred_result.scalars().all()
        rows = [(pred, games_by_id[pred.game_id]) for pred in preds if pred.game_id in games_by_id]

    await engine.dispose()

    if not rows:
        print(f"No predictions found for {today}. Exiting.")
        return

    print(f"Found {len(rows)} predictions for {today}:")
    for pred, game in rows:
        home = game.home_team.name if game.home_team else f"Team {game.home_team_id}"
        away = game.away_team.name if game.away_team else f"Team {game.away_team_id}"
        print(f"  {away} @ {home} | {game.commence_time} | status={game.status}")
        print(f"    fg_spread={pred.fg_spread} fg_total={pred.fg_total} ml={pred.fg_home_ml_prob}")

    # ── Prepare output paths ──────────────────────────────────────────────
    now_cst = datetime.now(_CST)
    stamp = now_cst.strftime("%H%M")
    date_str = today.strftime("%Y-%m-%d")

    date_dir = one_drive_root / date_str
    latest_dir = one_drive_root / "latest"
    date_dir.mkdir(parents=True, exist_ok=True)
    latest_dir.mkdir(parents=True, exist_ok=True)

    # ── Build CSV ─────────────────────────────────────────────────────────
    csv_content = build_slate_csv(rows, min_edge=0.0)  # include all picks
    csv_dated = date_dir / f"GBSV_NBA_Predictions_{stamp}.csv"
    csv_latest = latest_dir / "GBSV_NBA_Predictions.csv"
    csv_root = one_drive_root / f"nba_slate_{date_str}.csv"
    for output_path in [csv_dated, csv_latest, csv_root]:
        output_path.write_text(csv_content, encoding="utf-8")
    print(f"Wrote CSV: {csv_dated}")

    # ── Build HTML (interactive dark-themed slate) ────────────────────────
    html_content = build_html_slate(rows, odds_pulled_at=None)
    html_dated = date_dir / f"GBSV_NBA_Predictions_{stamp}.html"
    html_slate_dated = date_dir / f"GBSV_NBA_Slate_{stamp}.html"
    html_latest = latest_dir / "GBSV_NBA_Predictions.html"
    html_slate_latest = latest_dir / "GBSV_NBA_Slate.html"
    for output_path in [html_dated, html_slate_dated, html_latest, html_slate_latest]:
        output_path.write_text(html_content, encoding="utf-8")
    print(f"Wrote HTML (Predictions): {html_dated}")
    print(f"Wrote HTML (Slate):       {html_slate_dated}")

    # ── Build JSON ────────────────────────────────────────────────────────
    started_at = datetime.now(_CST)
    json_games: list[dict[str, Any]] = []
    for pred, game in rows:
        home = game.home_team.name if game.home_team else f"Team {game.home_team_id}"
        away = game.away_team.name if game.away_team else f"Team {game.away_team_id}"
        home_rec = _team_record(game.home_team) if game.home_team else ""
        away_rec = _team_record(game.away_team) if game.away_team else ""

        fg_spread = float(pred.fg_spread or 0)
        fg_total = float(pred.fg_total or 0)
        h1_spread = float(getattr(pred, "h1_spread", 0) or 0)
        h1_total = float(getattr(pred, "h1_total", 0) or 0)
        pred_home_fg = float(getattr(pred, "predicted_home_fg", 0) or 0)
        pred_away_fg = float(getattr(pred, "predicted_away_fg", 0) or 0)
        pred_home_1h = float(getattr(pred, "predicted_home_1h", 0) or 0)
        pred_away_1h = float(getattr(pred, "predicted_away_1h", 0) or 0)
        opening_spread = float(pred.opening_spread) if pred.opening_spread is not None else None
        opening_total = float(pred.opening_total) if pred.opening_total is not None else None

        odds_sourced = cast(dict[str, Any], getattr(pred, "odds_sourced", None) or {})
        books = cast(dict[str, dict[str, Any]], odds_sourced.get("books", {}))
        opening_h1_spread = odds_sourced.get("opening_h1_spread")
        opening_h1_total = odds_sourced.get("opening_h1_total")

        # Compute FG spread edge
        mkt_spread = opening_spread
        if mkt_spread is None and books:
            vals = [b["spread"] for b in books.values() if "spread" in b and b["spread"] is not None]
            mkt_spread = round(sum(vals) / len(vals), 1) if vals else None

        # Compute FG total edge
        mkt_total = opening_total
        if mkt_total is None and books:
            vals = [b["total"] for b in books.values() if "total" in b and b["total"] is not None]
            mkt_total = round(sum(vals) / len(vals), 1) if vals else None

        picks_list = extract_picks(pred, game, min_edge=0.0)

        # Build markets dict
        markets: dict[str, dict[str, Any]] = {}
        for pick in picks_list:
            key = f"{pick.segment}_{pick.market_type}"
            markets[key] = {
                "segment": pick.segment,
                "market": pick.market_type,
                "pick_label": pick.label,
                "line": pick.market_line,
                "edge": pick.edge,
                "fire_rating": pick.confidence,
                "odds": pick.odds,
                "rationale": pick.rationale,
            }

        ct_time = None
        if game.commence_time:
            naive_utc = game.commence_time.replace(tzinfo=None) if game.commence_time.tzinfo is None else game.commence_time
            if naive_utc.tzinfo is None:
                naive_utc = naive_utc.replace(tzinfo=UTC)
            ct_time = naive_utc.astimezone(_CST).isoformat()

        json_games.append({
            "event_id": str(game.odds_api_id or game.id),
            "home": home,
            "away": away,
            "game_time": ct_time,
            "home_record": home_rec,
            "away_record": away_rec,
            "status": game.status,
            "projected_scores": {
                "home_pred": round(pred_home_fg, 1),
                "away_pred": round(pred_away_fg, 1),
                "proj_total": round(fg_total, 1),
                "proj_margin": round(fg_spread, 1),
            },
            "projected_scores_1h": {
                "home_pred": round(pred_home_1h, 1),
                "away_pred": round(pred_away_1h, 1),
                "proj_total": round(h1_total, 1),
                "proj_margin": round(h1_spread, 1),
            },
            "market_lines": {
                "opening_spread": opening_spread,
                "opening_total": opening_total,
                "opening_h1_spread": opening_h1_spread,
                "opening_h1_total": opening_h1_total,
            },
            "picks": markets,
        })

    completed_at = datetime.now(_CST)
    json_payload = {
        "run_id": now_cst.strftime("%Y%m%d_%H%M%S"),
        "started_at": started_at.isoformat(),
        "started_at_display": started_at.strftime("%B %d, %Y %I:%M %p %Z"),
        "completed_at": completed_at.isoformat(),
        "completed_at_display": completed_at.strftime("%B %d, %Y %I:%M %p %Z"),
        "duration_seconds": (completed_at - started_at).total_seconds(),
        "games_found": len(rows),
        "games_predicted": len(rows),
        "model_version": str(getattr(rows[0][0], "model_version", "v6") if rows else "v6"),
        "date": date_str,
        "predictions": json_games,
        "errors": [],
    }
    json_content = json.dumps(json_payload, indent=2, default=str)
    json_dated = date_dir / f"GBSV_NBA_Predictions_{stamp}.json"
    json_latest = latest_dir / "GBSV_NBA_Predictions.json"
    for output_path in [json_dated, json_latest]:
        output_path.write_text(json_content, encoding="utf-8")
    print(f"Wrote JSON: {json_dated}")

    # ── Build TXT ─────────────────────────────────────────────────────────
    txt_lines = [
        f"GBSV NBA PREDICTIONS — {now_cst.strftime('%B %d, %Y %I:%M %p CT')}",
        f"Model: {json_payload['model_version']}",
        "=" * 60,
        "",
    ]
    for raw_game_data in json_games:
        game_data = cast(dict[str, Any], raw_game_data)
        proj = cast(dict[str, float], game_data["projected_scores"])
        proj1h = cast(dict[str, float], game_data["projected_scores_1h"])
        game_picks = cast(dict[str, dict[str, Any]], game_data.get("picks") or {})
        txt_lines.append(
            f"{game_data['away']} @ {game_data['home']}  |  {game_data.get('game_time', 'TBD')}"
        )
        txt_lines.append(
            f"  FG: {game_data['home']} {proj['home_pred']} – {game_data['away']} {proj['away_pred']}"
            f"  (spread {proj['proj_margin']:+.1f}, total {proj['proj_total']:.1f})"
        )
        txt_lines.append(
            f"  1H: {game_data['home']} {proj1h['home_pred']} – {game_data['away']} {proj1h['away_pred']}"
            f"  (spread {proj1h['proj_margin']:+.1f}, total {proj1h['proj_total']:.1f})"
        )
        if game_picks:
            txt_lines.append("  PICKS:")
            for raw_pick_data in game_picks.values():
                pk = cast(dict[str, Any], raw_pick_data)
                fires = "🔥" * pk["fire_rating"]
                txt_lines.append(
                    f"    [{pk['segment']} {pk['market']}] {pk['pick_label']}  "
                    f"edge={pk['edge']:.1f}  {fires}"
                )
                if pk.get("rationale"):
                    txt_lines.append(f"      {pk['rationale']}")
        txt_lines.append("")

    txt_lines += [
        "=" * 60,
        f"Generated by GBSV NBA v6 | {now_cst.strftime('%Y-%m-%d %H:%M CT')}",
    ]
    txt_content = "\n".join(txt_lines)
    txt_dated = date_dir / f"GBSV_NBA_Predictions_{stamp}.txt"
    txt_latest = latest_dir / "GBSV_NBA_Predictions.txt"
    for output_path in [txt_dated, txt_latest]:
        output_path.write_text(txt_content, encoding="utf-8")
    print(f"Wrote TXT: {txt_dated}")

    # ── Build XLSX ────────────────────────────────────────────────────────
    try:
        import csv as _csv
        import io as _io

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"NBA Picks {date_str}"

        # Parse lines from CSV to build XLSX
        csv_reader = _csv.DictReader(_io.StringIO(csv_content))
        headers = csv_reader.fieldnames or []

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        fire_colors = {5: "00B050", 4: "92D050", 3: "FFFF00", 2: "FFC000", 1: "FF0000"}

        for i, row_dict in enumerate(csv_reader, start=2):
            row_vals = [row_dict.get(h, "") for h in headers]
            ws.append(row_vals)
            try:
                edge_val = float(row_dict.get("Edge", 0) or 0)
                fire_cnt = (5 if edge_val >= 9 else 4 if edge_val >= 7 else
                            3 if edge_val >= 5 else 2 if edge_val >= 3.5 else 1)
                color = fire_colors.get(fire_cnt, "FFFFFF")
                for cell in ws[i]:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            except Exception:
                pass

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        xlsx_buf = _io.BytesIO()
        wb.save(xlsx_buf)
        xlsx_bytes = xlsx_buf.getvalue()

        xlsx_dated = date_dir / f"GBSV_NBA_Predictions_{stamp}.xlsx"
        xlsx_latest = latest_dir / "GBSV_NBA_Predictions.xlsx"
        for output_path in [xlsx_dated, xlsx_latest]:
            output_path.write_bytes(xlsx_bytes)
        print(f"Wrote XLSX: {xlsx_dated}")
    except ImportError:
        print("openpyxl not installed — skipping XLSX")

    print()
    print(f"Export complete -> {date_dir}")
    print(f"  CSV  : {csv_root.name}")
    print(f"  HTML : {html_dated.name}")
    print(f"  JSON : {json_dated.name}")
    print(f"  TXT  : {txt_dated.name}")


if __name__ == "__main__":
    asyncio.run(main())
